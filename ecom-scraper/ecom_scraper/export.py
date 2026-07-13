"""导出：多 Sheet Excel（明细 + 分类统计）。"""
from __future__ import annotations

import re
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(v):
    return _ILLEGAL.sub("", v) if isinstance(v, str) else v


def to_excel(items: list[dict], path: str, columns: list[str], group_by: str | None = None) -> None:
    import os
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    _header(ws, columns)
    for it in items:
        ws.append([_clean(it.get(c, "")) for c in columns])
    for ci, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(10, min(50, len(c) * 2 + 8))
    ws.freeze_panes = "A2"

    if group_by:
        ws2 = wb.create_sheet("分类统计")
        _header(ws2, [group_by, "数量"])
        for k, n in Counter(str(it.get(group_by, "")) for it in items).most_common():
            ws2.append([k, n])
    wb.save(path)


def _header(ws, cols: list[str]) -> None:
    ws.append(cols)
    fill = PatternFill("solid", fgColor="2E8B57")
    for ci, _ in enumerate(cols, 1):
        c = ws.cell(1, ci)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
